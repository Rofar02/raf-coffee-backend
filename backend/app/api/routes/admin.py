from pathlib import Path
import re
from typing import Any, Dict, List, Optional, Tuple
from io import BytesIO
import unicodedata
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from openpyxl import load_workbook
from app.core.database import get_pool
from app.repositories.dish_repository import DishRepository
from app.services.menu_service import MenuService
from app.repositories.vacancy_repository import VacancyRepository
from app.repositories.interior_repository import InteriorRepository
from app.services.vacancy_service import VacancyService
from app.models.dish import (
    CategoryItem,
    DishCreate,
    Dish,
    DishUpdate,
    MenuItem,
    Season,
    SeasonCreate,
    SeasonUpdate,
    SeasonDishLink,
    SubcategoryCreateBody,
    SubcategoryOption,
    SubcategoryUpdateBody,
)
from app.models.vacancy import (
    Vacancy,
    VacancyCreate,
    VacancyUpdate,
    VacancySettings,
    VacancySettingsUpdate,
    VacancyApplication,
)
from app.models.interior import InteriorGallery, InteriorGalleryUpdate
from app.core.auth import verify_admin_token

router = APIRouter(prefix="/admin", tags=["admin"])
_UPLOADS_DIR = Path("static") / "uploads"
_ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
_ALLOWED_VIDEO_TYPES = {"video/mp4", "video/webm", "video/quicktime"}
_EXT_BY_CONTENT_TYPE = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
    "video/mp4": ".mp4",
    "video/webm": ".webm",
    "video/quicktime": ".mov",
}

_CYR_TO_LAT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


class SeasonDishAttachPayload(BaseModel):
    sort_order: int = 0
    is_visible: bool = True


class XlsxImportResult(BaseModel):
    created: int
    updated: int
    skipped: int
    errors: List[str]
    warnings: List[str] = Field(default_factory=list)


def _slugify(value: str) -> str:
    raw = (value or "").strip().lower()
    translit = "".join(_CYR_TO_LAT.get(ch, ch) for ch in raw)
    normalized = unicodedata.normalize("NFKD", translit).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return slug or "season"


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", "_", text)
    return text


def _to_int(value: Any, *, field: str, row_num: int) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"row {row_num}: invalid int for '{field}' -> {value!r}")
    try:
        return int(value)
    except (TypeError, ValueError):
        text = str(value).strip()
        if not text:
            return None
        # Поддержка Excel-денег: "240.00", "240,00", "1 240,00 ₽"
        cleaned = re.sub(r"[^\d,.\-]", "", text).replace(",", ".")
        if cleaned.count(".") > 1:
            raise ValueError(f"row {row_num}: invalid int for '{field}' -> {value!r}")
        try:
            num = float(cleaned)
        except ValueError:
            raise ValueError(f"row {row_num}: invalid int for '{field}' -> {value!r}")
        if num.is_integer():
            return int(num)
        raise ValueError(f"row {row_num}: invalid int for '{field}' -> {value!r}")


def _to_bool(value: Any, *, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "да", "on"}


def _parse_volume_options(value: Any, *, row_num: int) -> List[Dict[str, int]]:
    if value is None:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    # Формат: "250:220;350:260" (объем:цена, через ; или ,)
    parts = [p.strip() for p in re.split(r"[;,]", raw) if p.strip()]
    out: List[Dict[str, int]] = []
    for i, part in enumerate(parts):
        if ":" not in part:
            raise ValueError(
                f"row {row_num}: invalid volume_options item '{part}', expected 'volume:price'"
            )
        left, right = [x.strip() for x in part.split(":", 1)]
        try:
            out.append({"volume_ml": int(left), "price": int(right), "sort_order": i})
        except (TypeError, ValueError):
            raise ValueError(
                f"row {row_num}: invalid volume_options numbers in '{part}', expected '250:220'"
            )
    return out


def _normalize_volume_text(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text:
        return None
    # Допускаем форматы: "250", "250 мл", "250ml", "250 г"
    m = re.search(r"\d+", text)
    if not m:
        return None
    try:
        return int(m.group(0))
    except (TypeError, ValueError):
        return None


# Типовые «верхние» разделы и подпись «Блюда» → в базе «Кухня» (и др. алиасы)
_HIERARCHICAL_BLOCK_NAMES = frozenset(
    {
        "блюда",
        "напитки",
        "кухня",
        "десерты",
        "завтраки",
        "закуски",
        "супы",
        "салаты",
        "второе",
        "мороженое",
    }
)
_HIERARCHICAL_CATEGORY_ALIASES: Dict[str, str] = {
    "блюда": "Кухня",
}


def _normalize_hierarchical_category_name(raw: str) -> str:
    t = (raw or "").strip()
    k = t.lower()
    if k in _HIERARCHICAL_CATEGORY_ALIASES:
        return _HIERARCHICAL_CATEGORY_ALIASES[k]
    return t


def _hierarchical_looks_like_section_row(
    *,
    name: str,
    volume_options: List[Dict[str, int]],
    explicit_price: Optional[int],
    has_any_meta: bool,
) -> bool:
    if not (name or "").strip():
        return False
    if volume_options:
        return False
    if explicit_price is not None:
        return False
    if has_any_meta:
        return False
    return True


def _name_column_cell_style(ws, row: int, col_1based: int):
    cell = ws.cell(row=row, column=col_1based)
    font = cell.font
    sz = None
    if font is not None and font.sz is not None:
        try:
            sz = float(font.sz)
        except (TypeError, ValueError):
            sz = None
    bold = bool(font and font.b)
    return sz, bold


def _merge_width_for_name_cell(ws, row: int, col_1based: int) -> int:
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= col_1based <= m.max_col:
            return int(m.max_col - m.min_col + 1)
    return 1


class _HierState:
    __slots__ = ("category", "subcategory", "had_dish_in_sub", "skipping_excluded_block")

    def __init__(self) -> None:
        self.category: Optional[str] = None
        self.subcategory: Optional[str] = None
        self.had_dish_in_sub: bool = False
        self.skipping_excluded_block: bool = False


def _is_excluded_hierarchical_block_title(raw: str) -> bool:
    """Заголовки, после которых в файле идут не блюда меню (сертификаты, розница) — не импортируем."""
    t = (raw or "").strip().lower()
    if not t:
        return False
    if "сертифик" in t:
        return True
    if "подароч" in t:
        return True
    if t == "розница" or t.startswith("розница ") or t.startswith("розница,"):
        return True
    return False


def _hierarchical_title_is_special_season(raw: str) -> bool:
    """
    «Спешл» / special — сезонные позиции: не в базовое меню, привязка к активному сезону при импорте.
    """
    t = (raw or "").strip().lower()
    if not t:
        return False
    if "спешл" in t:
        return True
    if "spesh" in t or "speshl" in t:
        return True
    if re.search(r"(?<![a-z])special(?![a-z])", t, re.IGNORECASE):
        return True
    return False


def _hier_special_scope(st: _HierState) -> bool:
    if _hierarchical_title_is_special_season(st.category or ""):
        return True
    if _hierarchical_title_is_special_season(st.subcategory or ""):
        return True
    return False


def _classify_hierarchical_section(
    st: _HierState,
    raw_title: str,
    *,
    font_sz: Optional[float],
    font_bold: bool,
    merge_cols: int,
) -> str:
    """
    «category» — крупный раздел (создаётся категория в БД);
    «subcategory» — подзаголовок (подкатегория в текущей категории).
    """
    t = (raw_title or "").strip()
    t_low = t.lower()
    is_block = t_low in _HIERARCHICAL_BLOCK_NAMES or t_low in _HIERARCHICAL_CATEGORY_ALIASES
    is_large = bool(
        merge_cols >= 4
        or (font_sz is not None and font_sz >= 14.0)
        or (font_sz is not None and font_sz >= 12.5 and font_bold)
    )
    if is_large:
        return "category"
    if st.category is None:
        return "category"
    if st.subcategory is None:
        if is_block:
            return "category"
        return "subcategory"
    if not st.had_dish_in_sub:
        if is_block:
            return "category"
        return "subcategory"
    if is_block:
        return "category"
    return "subcategory"


_TIER_ML = {2: 200, 3: 300, 4: 400}


def _parse_drink_tier_name(name: str) -> Optional[Tuple[str, int]]:
    """
    «Капучино 0.2» / «0,3» / «0.4» в конце названия -> базовое имя + объём в мл
    (0,2 = 200 мл, 0,3 = 300 мл, 0,4 = 400 мл).
    """
    raw = (name or "").strip()
    m = re.match(r"^(.+?)\s*0[.,](?P<d>[234])\s*$", raw, re.IGNORECASE)
    if not m:
        return None
    d = m.group("d")
    if d not in "234":
        return None
    return (m.group(1).strip(), _TIER_ML[int(d)])


def _hier_subcat_key(item: Dict[str, Any]) -> Tuple:
    return (
        item.get("subcategory_id"),
        item.get("category_name") or None,
        item.get("subcategory_name") or None,
    )


def _merge_tier_drink_prepared_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Склеивает «Капучино 0.2 / 0,3 / 0.4» в одно блюдо с volume_options (цены в колонке «Цена»)."""
    if not rows:
        return []
    out: List[Dict[str, Any]] = []
    i = 0
    while i < len(rows):
        a = rows[i]
        t1 = _parse_drink_tier_name(str(a.get("name") or ""))
        if not t1:
            out.append(a)
            i += 1
            continue
        base, mla = t1
        key_a = _hier_subcat_key(a)
        group: List[Dict[str, Any]] = [a]
        mls = {mla}
        j = i + 1
        while j < len(rows):
            b = rows[j]
            t2 = _parse_drink_tier_name(str(b.get("name") or ""))
            if not t2 or t2[0] != base or _hier_subcat_key(b) != key_a:
                break
            mlb = t2[1]
            if mlb in mls:
                break
            group.append(b)
            mls.add(mlb)
            j += 1

        with_vol: List[Tuple[int, Dict[str, Any]]] = []
        for g in group:
            pr = _parse_drink_tier_name(str(g.get("name") or ""))
            if not pr:
                with_vol = []
                break
            with_vol.append((pr[1], g))
        with_vol.sort(key=lambda x: x[0])
        if not with_vol:
            out.append(a)
            i += 1
            continue

        g0m = with_vol[0][1]
        more_than_one = len(with_vol) > 1

        vopts2: List[Dict[str, Any]] = []
        for idx, (vml, gr) in enumerate(with_vol):
            p = int(gr.get("explicit_price") or 0)
            kc = str(gr.get("kcal_text") or "").strip() or None
            bj = str(gr.get("bju_text") or "").strip() or None
            opt: Dict[str, Any] = {
                "volume_ml": vml,
                "price": p,
                "sort_order": idx,
            }
            if kc:
                opt["nutrition_kcal"] = kc
            if bj:
                opt["nutrition_bju"] = bj
            vopts2.append(opt)

        out.append(
            {
                "row_num": g0m.get("row_num"),
                "name": base,
                "dish_id_from_file": g0m.get("dish_id_from_file"),
                "explicit_price": min(int(x[1].get("explicit_price") or 0) for x in with_vol),
                "volume_options": vopts2,
                "weight_grams": None if more_than_one else g0m.get("weight_grams"),
                "volume_ml": None,
                "description": next(
                    (
                        str(x[1].get("description") or "").strip()
                        for x in with_vol
                        if str(x[1].get("description") or "").strip()
                    ),
                    g0m.get("description"),
                ),
                "calories": None,
                "image_url": g0m.get("image_url"),
                "video_url": g0m.get("video_url"),
                "is_base_menu": g0m.get("is_base_menu", True),
                "subcategory_id": g0m.get("subcategory_id"),
                "category_name": g0m.get("category_name"),
                "subcategory_name": g0m.get("subcategory_name"),
                "hierarchy_from_rows": g0m.get("hierarchy_from_rows", False),
                "hier_special": bool(g0m.get("hier_special")),
                "tier_merged": True,
            }
        )
        i = j

    for row in out:
        row.pop("kcal_text", None)
        row.pop("bju_text", None)
    return out


@router.get("/subcategories", response_model=List[SubcategoryOption])
async def list_subcategories(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    return await repo.list_subcategories()


@router.get("/categories", response_model=List[CategoryItem])
async def list_categories(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    rows = await repo.list_categories()
    return [CategoryItem(**r) for r in rows]


@router.post("/flush-menu-cache")
async def admin_flush_menu_cache(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    """Сбрасывает Redis-кеш публичного меню. Нужен после ручного SQL в PostgreSQL (кэш иначе до 5 мин)."""
    repo = DishRepository(pool)
    service = MenuService(repo)
    await service.clear_menu_cache()
    return {"ok": True}


@router.post("/subcategories", response_model=SubcategoryOption)
async def create_subcategory(
    body: SubcategoryCreateBody,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    cat = await repo.get_category_by_id(body.category_id)
    if not cat:
        raise HTTPException(status_code=400, detail="Category not found")
    row = await repo.create_subcategory(body.category_id, body.name, body.sort_order)
    return SubcategoryOption(**row)


@router.put("/subcategories/{subcategory_id}", response_model=SubcategoryOption)
async def update_subcategory(
    subcategory_id: int,
    body: SubcategoryUpdateBody,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    data = body.model_dump(exclude_none=True)
    if not data:
        raise HTTPException(status_code=400, detail="No fields to update")
    if data.get("category_id") is not None:
        cat = await repo.get_category_by_id(int(data["category_id"]))
        if not cat:
            raise HTTPException(status_code=400, detail="Category not found")
    ok = await repo.update_subcategory(
        subcategory_id,
        data.get("category_id"),
        data.get("name"),
        data.get("sort_order"),
    )
    if not ok:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    out = await repo.get_subcategory_by_id(subcategory_id)
    if not out:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    return SubcategoryOption(**out)


@router.delete("/subcategories/{subcategory_id}")
async def delete_subcategory(
    subcategory_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    deleted = await repo.delete_subcategory(subcategory_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    return {"ok": True, "deleted_id": subcategory_id}


@router.post("/dishes", response_model=MenuItem)
async def add_dish(
    dish: DishCreate,
    season_id: int | None = None,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    data = dish.model_dump()
    dish_id = await service.add_dish(data)
    if season_id is not None:
        await service.attach_dish_to_season(season_id=season_id, dish_id=dish_id)
    out = await repo.get_dish_by_id(dish_id)
    if not out:
        raise HTTPException(status_code=500, detail="Dish not found after create")
    return MenuItem(**out)


@router.get("/dishes", response_model=List[MenuItem])
async def list_dishes(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    return await repo.get_all()


@router.put("/dishes/{dish_id}", response_model=MenuItem)
async def update_dish(
    dish_id: int,
    dish: DishUpdate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    data = dish.model_dump()
    updated = await service.update_dish(dish_id, data)
    if not updated:
        raise HTTPException(status_code=404, detail="Dish not found")
    out = await repo.get_dish_by_id(dish_id)
    if not out:
        raise HTTPException(status_code=404, detail="Dish not found")
    return MenuItem(**out)


@router.delete("/dishes/{dish_id}")
async def delete_dish(
    dish_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    deleted = await service.delete_dish(dish_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Dish not found")
    return {"ok": True, "deleted_id": dish_id}


@router.get("/seasons", response_model=List[Season])
async def list_seasons(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    return await service.list_seasons()


@router.post("/seasons", response_model=Season)
async def create_season(
    season: SeasonCreate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    payload = season.model_dump()
    existing = await service.list_seasons()
    existing_slugs = {str(row.get("slug") or "").strip().lower() for row in existing}

    base_slug = _slugify(payload.get("slug") or payload.get("name") or "season")
    final_slug = base_slug
    idx = 2
    while final_slug in existing_slugs:
        final_slug = f"{base_slug}-{idx}"
        idx += 1
    payload["slug"] = final_slug

    created = await service.create_season(payload)
    return Season(**created)


@router.put("/seasons/{season_id}", response_model=Season)
async def update_season(
    season_id: int,
    season: SeasonUpdate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    updated = await service.update_season(season_id, season.model_dump(exclude_none=True))
    if not updated:
        raise HTTPException(status_code=404, detail="Season not found")
    return Season(**updated)


@router.post("/seasons/{season_id}/activate")
async def activate_season(
    season_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    ok = await service.set_active_season(season_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Season not found")
    return {"ok": True, "active_season_id": season_id}


@router.delete("/seasons/{season_id}")
async def delete_season(
    season_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    result = await service.delete_season(season_id)
    if not result.get("ok"):
        reason = result.get("reason")
        if reason == "last_season":
            raise HTTPException(status_code=400, detail="Cannot delete last season")
        raise HTTPException(status_code=404, detail="Season not found")
    return {"ok": True, "deleted_season_id": season_id}


@router.get("/seasons/{season_id}/dishes", response_model=List[MenuItem])
async def list_season_dishes(
    season_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    return await service.list_dishes_by_season(season_id)


@router.post("/seasons/{season_id}/dishes/{dish_id}", response_model=SeasonDishLink)
async def attach_dish_to_season(
    season_id: int,
    dish_id: int,
    payload: SeasonDishAttachPayload = SeasonDishAttachPayload(),
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    ok = await service.attach_dish_to_season(
        season_id=season_id,
        dish_id=dish_id,
        sort_order=payload.sort_order,
        is_visible=payload.is_visible,
    )
    if not ok:
        raise HTTPException(status_code=404, detail="Season or dish not found")
    return SeasonDishLink(
        season_id=season_id,
        dish_id=dish_id,
        sort_order=payload.sort_order,
        is_visible=payload.is_visible,
    )


@router.delete("/seasons/{season_id}/dishes/{dish_id}")
async def detach_dish_from_season(
    season_id: int,
    dish_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = DishRepository(pool)
    service = MenuService(repo)
    ok = await service.detach_dish_from_season(season_id=season_id, dish_id=dish_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Season-dish link not found")
    return {"ok": True, "season_id": season_id, "dish_id": dish_id}


@router.post("/upload")
async def upload_media(
    file: UploadFile = File(...),
    _: str = Depends(verify_admin_token),
):
    content_type = (file.content_type or "").lower()
    if content_type in _ALLOWED_IMAGE_TYPES:
        kind = "images"
        ext = _EXT_BY_CONTENT_TYPE[content_type]
    elif content_type in _ALLOWED_VIDEO_TYPES:
        kind = "videos"
        ext = _EXT_BY_CONTENT_TYPE[content_type]
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    target_dir = _UPLOADS_DIR / kind
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid4().hex}{ext}"
    target_path = target_dir / filename

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(payload) > 60 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large")

    target_path.write_bytes(payload)
    return {"url": f"/static/uploads/{kind}/{filename}", "content_type": content_type}


@router.post("/import-xlsx", response_model=XlsxImportResult)
async def import_menu_from_xlsx(
    file: UploadFile = File(...),
    strict: bool = Query(True, description="If true, rollback-like behavior: no writes when any row has errors"),
    hierarchical: bool = Query(
        False,
        description="Категория/подкатегория в виде «блоков» в колонке названия: крупные и мелкие заголовки, затем строки с ценой",
    ),
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")
    ws = None
    for candidate in wb.worksheets:
        if candidate.max_row <= 0 or candidate.max_column <= 0:
            continue
        has_any = False
        for r in range(1, min(candidate.max_row, 20) + 1):
            row_vals = [candidate.cell(row=r, column=c).value for c in range(1, candidate.max_column + 1)]
            if any(str(v).strip() for v in row_vals if v is not None):
                has_any = True
                break
        if has_any:
            ws = candidate
            break
    if ws is None:
        raise HTTPException(status_code=400, detail="XLSX file has no readable rows")

    header_row_idx = None
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v is not None and str(v).strip()]
        if len(non_empty) >= 2:
            header_row_idx = r
            break
    if header_row_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Could not detect header row. Put headers in one row without merged-only title rows.",
        )

    if ws.max_row < header_row_idx + 1:
        raise HTTPException(status_code=400, detail="XLSX must contain at least one data row below headers")

    header_cells = [cell.value for cell in ws[header_row_idx]]
    header_map: Dict[str, int] = {}
    for idx, raw in enumerate(header_cells):
        norm = _norm_header(raw)
        if norm:
            header_map[norm] = idx

    aliases = {
        "id": ["id", "dish_id", "dishid"],
        "name": [
            "name",
            "название",
            "наименование",
            "наименование_товара",
            "наименование_блюда",
            "товар",
            "позиция",
            "dish_name",
        ],
        "price": ["price", "цена", "стоимость"],
        "weight_grams": [
            "weight_grams",
            "weight",
            "вес",
            "вес_г",
            "вес,_г/объем,_мл",
            "вес_г/объем_мл",
        ],
        "volume_ml": [
            "volume_ml",
            "volume",
            "объем",
            "объем_мл",
            "объём",
            "объём_мл",
            "вес_-_объем",
            "вес_-_объём",
            "вес-объем",
            "вес-объём",
            "вес/объем",
            "вес/объём",
            "вес_объем",
            "вес_объём",
        ],
        "description": ["description", "описание", "состав"],
        "calories": ["calories", "ккал", "калории", "калорийность"],
        "bju": ["бжус", "бжу", "bju", "protein_fat_carb"],
        "image_url": ["image_url", "image", "фото", "картинка"],
        "video_url": ["video_url", "video", "видео"],
        "is_base_menu": ["is_base_menu", "base_menu", "базовое_меню"],
        "subcategory_id": ["subcategory_id", "subcat_id"],
        "category_name": ["category", "category_name", "категория"],
        "subcategory_name": ["subcategory", "subcategory_name", "подкатегория"],
        "volume_options": ["volume_options", "объемы", "объёмы", "volume_prices"],
    }

    cols: Dict[str, int] = {}
    for field, keys in aliases.items():
        for key in keys:
            if key in header_map:
                cols[field] = header_map[key]
                break

    if "name" not in cols:
        detected = sorted(header_map.keys())
        raise HTTPException(
            status_code=400,
            detail={
                "message": "XLSX must contain a dish name column",
                "accepted": ["name", "название", "наименование", "наименование товара"],
                "detected_headers": detected,
            },
        )

    repo = DishRepository(pool)
    service = MenuService(repo)

    result: Dict[str, Any] = {"created": 0, "updated": 0, "skipped": 0, "errors": [], "warnings": []}
    prepared_rows: List[Dict[str, Any]] = []
    hier_st: Optional[_HierState] = _HierState() if hierarchical else None

    for row_num in range(header_row_idx + 1, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_num]]
        row_data: Dict[str, Any] = {}
        for field, col_idx in cols.items():
            row_data[field] = row_values[col_idx] if col_idx < len(row_values) else None

        name = str(row_data.get("name") or "").strip()
        if not name:
            result["skipped"] += 1
            continue

        try:
            if hierarchical:
                subcategory_id = None
            else:
                subcategory_id = _to_int(
                    row_data.get("subcategory_id"), field="subcategory_id", row_num=row_num
                )
            if not hierarchical:
                category_name = str(row_data.get("category_name") or "").strip()
                subcategory_name = str(row_data.get("subcategory_name") or "").strip()
                if (category_name and not subcategory_name) or (subcategory_name and not category_name):
                    raise ValueError(
                        f"row {row_num}: both 'category' and 'subcategory' must be set together"
                    )
            else:
                category_name = ""
                subcategory_name = ""

            volume_options = (
                _parse_volume_options(row_data.get("volume_options"), row_num=row_num)
                if "volume_options" in cols
                else []
            )
            explicit_price = (
                _to_int(row_data.get("price"), field="price", row_num=row_num)
                if "price" in cols
                else None
            )
            has_any_meta = any(
                [
                    str(row_data.get("volume_ml") or "").strip(),
                    str(row_data.get("description") or "").strip(),
                    str(row_data.get("calories") or "").strip(),
                    str(row_data.get("bju") or "").strip(),
                    str(row_data.get("category_name") or "").strip(),
                    str(row_data.get("subcategory_name") or "").strip(),
                    str(row_data.get("weight_grams") or "").strip(),
                ]
            )
            if hierarchical and hier_st is not None and _hierarchical_looks_like_section_row(
                name=name,
                volume_options=volume_options,
                explicit_price=explicit_price,
                has_any_meta=has_any_meta,
            ):
                if _is_excluded_hierarchical_block_title(name):
                    hier_st.skipping_excluded_block = True
                    result["skipped"] += 1
                    continue
                hier_st.skipping_excluded_block = False
                name_col_1b = int(cols["name"]) + 1
                f_sz, f_bold = _name_column_cell_style(ws, row_num, name_col_1b)
                merge_w = _merge_width_for_name_cell(ws, row_num, name_col_1b)
                _kind = _classify_hierarchical_section(
                    hier_st, name, font_sz=f_sz, font_bold=f_bold, merge_cols=merge_w
                )
                if _kind == "category":
                    hier_st.category = _normalize_hierarchical_category_name(name)
                    hier_st.subcategory = None
                    hier_st.had_dish_in_sub = False
                else:
                    if not hier_st.category:
                        raise ValueError(
                            f"row {row_num}: подкатегория «{name}» идёт до первого крупного раздела (категории)"
                        )
                    hier_st.subcategory = name.strip()
                    hier_st.had_dish_in_sub = False
                continue

            if hierarchical and hier_st is not None and hier_st.skipping_excluded_block:
                result["skipped"] += 1
                continue

            if hierarchical and hier_st is not None:
                if not hier_st.category or not hier_st.subcategory:
                    raise ValueError(
                        f"row {row_num}: в иерархическом XLSX сначала укажите крупный раздел и подраздел, затем строки с ценой (или ID)"
                    )
                category_name = hier_st.category
                subcategory_name = hier_st.subcategory
                subcategory_id = None
            if not volume_options and explicit_price is None:
                if not has_any_meta:
                    # Часто в прайсах есть строки-разделители (только название без цены).
                    result["skipped"] += 1
                    continue
                raise ValueError(
                    f"row {row_num}: 'price' is required when 'volume_options' is empty"
                )

            weight_grams = None
            if "weight_grams" in cols:
                weight_grams = _to_int(
                    row_data.get("weight_grams"), field="weight_grams", row_num=row_num
                )
            category_hint = (category_name or "").strip().lower()
            is_drink_hint = "напит" in category_hint
            if weight_grams is None and "volume_ml" in cols and not is_drink_hint:
                # Для кухонных позиций из объединенной колонки берем значение как вес (г).
                weight_grams = _normalize_volume_text(row_data.get("volume_ml"))
            volume_ml = None
            if "volume_ml" in cols:
                raw_volume = row_data.get("volume_ml")
                try:
                    volume_ml = _to_int(raw_volume, field="volume_ml", row_num=row_num)
                except ValueError:
                    volume_ml = _normalize_volume_text(raw_volume)
            if volume_ml is None and "volume_ml" in cols and is_drink_hint:
                # Для напитков из объединенной колонки берем значение как объем (мл).
                volume_ml = _normalize_volume_text(row_data.get("volume_ml"))
            if _parse_drink_tier_name(name):
                # 0,2/0,3/0,4 в названии — объём в мл, не путать с весом 211/314 г в колонке.
                volume_ml = None
            description = (
                str(row_data.get("description") or "").strip() or None
                if "description" in cols
                else None
            )
            kcal_text = (
                str(row_data.get("calories") or "").strip() or None
                if "calories" in cols
                else None
            )
            bju_text = (
                str(row_data.get("bju") or "").strip() or None
                if "bju" in cols
                else None
            )
            if bju_text and kcal_text:
                calories = f"{kcal_text}\n{bju_text}"
            elif bju_text:
                calories = bju_text
            else:
                calories = kcal_text
            image_url = (
                str(row_data.get("image_url") or "").strip() or None
                if "image_url" in cols
                else None
            )
            video_url = (
                str(row_data.get("video_url") or "").strip() or None
                if "video_url" in cols
                else None
            )
            is_base_menu = (
                _to_bool(row_data.get("is_base_menu"), default=True)
                if "is_base_menu" in cols
                else True
            )
            hier_special = bool(hier_st and hierarchical and _hier_special_scope(hier_st))
            if hier_special:
                is_base_menu = False
            dish_id_from_file = _to_int(row_data.get("id"), field="id", row_num=row_num)

            prepared_rows.append(
                {
                    "row_num": row_num,
                    "name": name,
                    "dish_id_from_file": dish_id_from_file,
                    "explicit_price": explicit_price,
                    "volume_options": volume_options,
                    "weight_grams": weight_grams,
                    "volume_ml": volume_ml,
                    "description": description,
                    "calories": calories,
                    "kcal_text": kcal_text,
                    "bju_text": bju_text,
                    "image_url": image_url,
                    "video_url": video_url,
                    "is_base_menu": is_base_menu,
                    "subcategory_id": subcategory_id,
                    "category_name": category_name,
                    "subcategory_name": subcategory_name,
                    "hierarchy_from_rows": bool(hierarchical),
                    "hier_special": hier_special,
                }
            )
            if hier_st is not None:
                hier_st.had_dish_in_sub = True
        except Exception as e:
            result["errors"].append(str(e))

    prepared_rows = _merge_tier_drink_prepared_rows(prepared_rows)

    if strict and result["errors"]:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Import failed in strict mode: no changes were applied",
                "errors": result["errors"],
            },
        )

    seasons = await service.list_seasons()
    active_season = next((s for s in seasons if s.get("is_active")), None)
    active_season_id: Optional[int] = int(active_season["id"]) if active_season else None
    warned_spehsl_no_season = False

    for item in prepared_rows:
        row_num = item["row_num"]
        try:
            subcategory_id = item["subcategory_id"]
            category_name = item["category_name"]
            subcategory_name = item["subcategory_name"]
            if subcategory_id is None and category_name and subcategory_name:
                category = await repo.get_category_by_name(category_name)
                if not category:
                    category = await repo.create_category(category_name, sort_order=0)
                if not category:
                    raise ValueError(f"row {row_num}: cannot create category '{category_name}'")
                subcat = await repo.get_subcategory_by_name(category["id"], subcategory_name)
                if not subcat:
                    subcat = await repo.create_subcategory(category["id"], subcategory_name, sort_order=0)
                subcategory_id = int(subcat["id"])

            target_dish = None
            if item["dish_id_from_file"] is not None:
                target_dish = await repo.get_dish_by_id(item["dish_id_from_file"])
            if not target_dish:
                target_dish = await repo.find_dish_by_name_and_subcategory(item["name"], subcategory_id)

            if item["volume_options"]:
                final_price = min(v["price"] for v in item["volume_options"])
            elif item["explicit_price"] is not None:
                final_price = int(item["explicit_price"])
            elif target_dish:
                final_price = int(target_dish.get("price") or 0)
            else:
                raise ValueError(f"row {row_num}: cannot determine final price")

            dish_payload: Dict[str, Any] = {
                "name": item["name"],
                "price": final_price,
                "weight_grams": item["weight_grams"],
                "volume_ml": item["volume_ml"],
                "description": item["description"],
                "calories": item["calories"],
                "image_url": item["image_url"],
                "video_url": item["video_url"],
                "is_base_menu": item["is_base_menu"],
                "subcategory_id": subcategory_id,
                "volume_options": item["volume_options"] or None,
            }

            if target_dish:
                merged = {
                    "name": dish_payload["name"] or target_dish.get("name"),
                    "price": dish_payload["price"] if ("price" in cols or "volume_options" in cols) else int(target_dish.get("price") or 0),
                    "weight_grams": dish_payload["weight_grams"] if "weight_grams" in cols else target_dish.get("weight_grams"),
                    "volume_ml": dish_payload["volume_ml"] if "volume_ml" in cols else target_dish.get("volume_ml"),
                    "description": dish_payload["description"] if "description" in cols else target_dish.get("description"),
                    "calories": dish_payload["calories"] if "calories" in cols else target_dish.get("calories"),
                    "image_url": dish_payload["image_url"] if "image_url" in cols else target_dish.get("image_url"),
                    "video_url": dish_payload["video_url"] if "video_url" in cols else target_dish.get("video_url"),
                    "is_base_menu": (
                        dish_payload["is_base_menu"]
                        if ("is_base_menu" in cols or item.get("hier_special"))
                        else bool(target_dish.get("is_base_menu", True))
                    ),
                    "subcategory_id": dish_payload["subcategory_id"]
                    if (
                        "subcategory_id" in cols
                        or ("category_name" in cols and "subcategory_name" in cols)
                        or item.get("hierarchy_from_rows")
                    )
                    else target_dish.get("subcategory_id"),
                    "volume_options": dish_payload["volume_options"]
                    if ("volume_options" in cols or item.get("tier_merged"))
                    else target_dish.get("volume_options"),
                }
                ok = await service.update_dish(int(target_dish["id"]), merged)
                if ok:
                    result["updated"] += 1
                    if item.get("hier_special") and active_season_id is not None:
                        await service.attach_dish_to_season(
                            active_season_id,
                            int(target_dish["id"]),
                            sort_order=int(row_num),
                            is_visible=True,
                        )
                    elif item.get("hier_special") and not warned_spehsl_no_season:
                        result["warnings"].append(
                            "В прайсе есть «спешл» (сезон), но в базе нет активного сезона: блюда не попали в сезон, только is_base_menu=false. Создайте сезон и сделайте его активным, затем повторите импорт или привяжите блюда вручную."
                        )
                        warned_spehsl_no_season = True
                else:
                    result["errors"].append(f"row {row_num}: failed to update dish id={target_dish['id']}")
            else:
                new_id = await service.add_dish(dish_payload)
                result["created"] += 1
                if item.get("hier_special") and active_season_id is not None:
                    await service.attach_dish_to_season(
                        active_season_id,
                        int(new_id),
                        sort_order=int(row_num),
                        is_visible=True,
                    )
                elif item.get("hier_special") and not warned_spehsl_no_season:
                    result["warnings"].append(
                        "В прайсе есть «спешл» (сезон), но в базе нет активного сезона: блюда не попали в сезон, только is_base_menu=false. Создайте сезон и сделайте его активным, затем повторите импорт или привяжите блюда вручную."
                    )
                    warned_spehsl_no_season = True
        except Exception as e:
            result["errors"].append(str(e))
            if strict:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Import aborted in strict mode",
                        "errors": result["errors"],
                    },
                )

    await service.clear_menu_cache()
    return XlsxImportResult(**result)


@router.get("/vacancies/settings", response_model=VacancySettings)
async def get_vacancy_settings(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    payload = await service.get_settings()
    return VacancySettings(**payload)


@router.put("/vacancies/settings", response_model=VacancySettings)
async def update_vacancy_settings(
    body: VacancySettingsUpdate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    payload = await service.update_settings(show_on_homepage=body.show_on_homepage)
    return VacancySettings(**payload)


@router.get("/vacancies", response_model=List[Vacancy])
async def list_vacancies(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    rows = await service.list_vacancies()
    return [Vacancy(**x) for x in rows]


@router.post("/vacancies", response_model=Vacancy)
async def create_vacancy(
    body: VacancyCreate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    row = await service.create_vacancy(body.model_dump())
    return Vacancy(**row)


@router.put("/vacancies/{vacancy_id}", response_model=Vacancy)
async def update_vacancy(
    vacancy_id: int,
    body: VacancyUpdate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    row = await service.update_vacancy(vacancy_id, body.model_dump(exclude_unset=True))
    if not row:
        raise HTTPException(status_code=404, detail="Vacancy not found")
    return Vacancy(**row)


@router.delete("/vacancies/{vacancy_id}")
async def delete_vacancy(
    vacancy_id: int,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    ok = await service.delete_vacancy(vacancy_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Vacancy not found")
    return {"ok": True, "deleted_id": vacancy_id}


@router.get("/vacancy-applications", response_model=List[VacancyApplication])
async def list_vacancy_applications(
    limit: int = Query(200, ge=1, le=1000),
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = VacancyRepository(pool)
    service = VacancyService(repo)
    rows = await service.list_applications(limit=limit)
    return [VacancyApplication(**x) for x in rows]


@router.get("/interior", response_model=List[InteriorGallery])
async def list_interior_galleries(
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = InteriorRepository(pool)
    rows = await repo.list_all()
    return [InteriorGallery(**x) for x in rows]


@router.put("/interior/{slug}", response_model=InteriorGallery)
async def update_interior_gallery(
    slug: str,
    body: InteriorGalleryUpdate,
    _: str = Depends(verify_admin_token),
    pool=Depends(get_pool),
):
    repo = InteriorRepository(pool)
    row = await repo.upsert(slug=slug, payload=body.model_dump(exclude_unset=True))
    return InteriorGallery(**row)